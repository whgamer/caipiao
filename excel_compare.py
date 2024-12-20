import pandas as pd
from tkinter import Tk, filedialog, Button, Label, messagebox
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def load_excel(file_path):
    try:
        return pd.read_excel(file_path)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load file: {e}")
        return None

def compare_excel(file1, file2):
    df1 = load_excel(file1)
    df2 = load_excel(file2)

    if df1 is None or df2 is None:
        return None, None, None

    common_columns = set(df1.columns).intersection(set(df2.columns))
    unique_in_file1 = set(df1.columns) - common_columns
    unique_in_file2 = set(df2.columns) - common_columns

    comparison_results = {}
    for column in common_columns:
        differences = df1[column].fillna("NULL") != df2[column].fillna("NULL")
        comparison_results[column] = differences

    return unique_in_file1, unique_in_file2, comparison_results

def save_with_colors(file1, file2, output_path, comparison_results):
    df1 = load_excel(file1)
    df2 = load_excel(file2)
    
    wb = Workbook()
    ws = wb.active

    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    fill_green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    ws.append(["Column"] + list(df1.columns))

    for column, differences in comparison_results.items():
        for i, is_different in enumerate(differences):
            value1 = df1[column].iloc[i] if i < len(df1) else ""
            value2 = df2[column].iloc[i] if i < len(df2) else ""
            color = fill_red if is_different else fill_green
            ws.cell(row=i+2, column=1).value = column
            ws.cell(row=i+2, column=2).value = value1
            ws.cell(row=i+2, column=3).value = value2
            ws.cell(row=i+2, column=1).fill = color

    wb.save(output_path)
    messagebox.showinfo("Success", f"Comparison saved to {output_path}")

def select_files():
    file1 = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if not file1:
        return None

    file2 = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if not file2:
        return None

    return file1, file2

def compare_and_save():
    files = select_files()
    if not files:
        return

    file1, file2 = files
    unique_in_file1, unique_in_file2, comparison_results = compare_excel(file1, file2)

    if unique_in_file1 or unique_in_file2:
        unique_message = f"Unique columns in File 1: {unique_in_file1}\nUnique columns in File 2: {unique_in_file2}"
        messagebox.showinfo("Unique Columns", unique_message)

    output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
    if not output_path:
        return

    save_with_colors(file1, file2, output_path, comparison_results)

def create_gui():
    root = Tk()
    root.title("Excel Comparison Tool")

    Label(root, text="Excel Comparison Tool", font=("Arial", 14)).pack(pady=10)

    compare_button = Button(root, text="Compare Excel Files", command=compare_and_save, width=30)
    compare_button.pack(pady=20)

    root.mainloop()

if __name__ == "__main__":
    create_gui()
